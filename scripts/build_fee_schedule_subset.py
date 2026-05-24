import json
import sys
from pathlib import Path

from openpyxl import load_workbook


RAW_DIR = Path(sys.argv[1] if len(sys.argv) > 1 else "data/raw")
OUTPUT = Path(sys.argv[2] if len(sys.argv) > 2 else "frontend/assets/js/fee_schedule_items.js")

PREFIXES = set("DEFGHJKLMNPQRSTUVWXY")


def to_int(value):
    if value in (None, ""):
        return 0
    return int(float(str(value).replace(",", "").strip() or 0))


def infer_group(code, name, is_surgery):
    text = name.lower()
    if is_surgery or code[:1] in set("MNOPQRS") or any(word in name for word in ["수술", "절제", "봉합", "절개술"]):
        return "surgery"
    if code[:1] in set("LKJU") or any(word in name for word in ["마취", "주사", "처치", "치료", "재활"]):
        return "procedure_hira"
    if code[:1] in set("DCEFGH") or any(word.lower() in text for word in ["검사", "ct", "mri", "촬영", "초음파", "내시경"]):
        return "test"
    return "etc"


def infer_category(code, group, name):
    if group == "surgery":
        return "surgery"
    if group == "procedure_hira":
        return "procedure"
    if code[:1] in set("GH") or any(word in name for word in ["촬영", "조영", "초음파", "자기공명", "전산화"]):
        return "imaging"
    if code[:1] == "D" or "혈액" in name or "소변" in name:
        return "specimen"
    if "내시경" in name:
        return "endoscopy"
    return "functional"


def build_keywords(code, name, group):
    parts = {code.lower(), name.lower()}
    for token in ["ct", "mri", "x-ray", "엑스레이", "초음파", "내시경", "혈액", "소변", "마취", "주사", "처치", "수술"]:
        if token in name.lower():
            parts.add(token)
    if group == "test":
        parts.add("검사")
    elif group == "procedure_hira":
        parts.add("시술")
    elif group == "surgery":
        parts.add("수술")
    return sorted(parts)


def main():
    xlsx_files = sorted(RAW_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not xlsx_files:
        raise SystemExit("No xlsx file found in data/raw")

    workbook = load_workbook(xlsx_files[0], read_only=True, data_only=True)
    items = []

    for sheet_index, benefit_type in [(0, "benefit"), (1, "non-benefit"), (2, "full-self-pay")]:
        ws = workbook.worksheets[sheet_index]
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[0] or "").strip()
            name = str(row[3] or "").strip()
            if not code or not name:
                continue
            if code[:1] not in PREFIXES or len(code) > 5:
                continue

            clinic_price = to_int(row[7])
            hospital_price = to_int(row[8])
            if clinic_price <= 0 and hospital_price <= 0:
                continue

            is_surgery = str(row[6] or "").strip() == "1"
            group = infer_group(code, name, is_surgery)
            category = infer_category(code, group, name)

            items.append(
                {
                    "code": f"FEE_{code}",
                    "publicActionCode": code,
                    "category": category,
                    "group": group,
                    "type": code[:2],
                    "name": name,
                    "price": hospital_price or clinic_price,
                    "clinicPrice": clinic_price,
                    "hospitalPrice": hospital_price or clinic_price,
                    "isBenefit": benefit_type == "benefit",
                    "isDRG": False,
                    "isSurgery": is_surgery,
                    "alreadyPricedByProvider": True,
                    "publicFeeScheduleSource": f"{workbook.worksheets[sheet_index].title} {code}",
                    "keywords": build_keywords(code, name, group),
                }
            )

    payload = {
        "sourceFile": str(xlsx_files[0]).replace("\\", "/"),
        "sourceDate": "2026-05-01",
        "count": len(items),
        "items": items,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        "window.PUBLIC_FEE_SCHEDULE_ITEMS = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"Generated {OUTPUT} ({len(items)} items)")


if __name__ == "__main__":
    main()
